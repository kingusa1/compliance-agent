"""Extract Phase-2 XLSX reference files (rejection lists + tracker example)
into plain markdown so the orchestrator and any subagent can read them
without parsing spreadsheets.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "compliance-docs" / "COMPLIANCE XAI"
OUT = ROOT / ".planning" / "phase2-docs"
OUT.mkdir(parents=True, exist_ok=True)


def xlsx_to_md(path: Path) -> str:
    wb = load_workbook(path, data_only=True, read_only=True)
    parts: list[str] = [f"# {path.stem}\n"]
    for ws in wb.worksheets:
        parts.append(f"\n## Sheet: {ws.title}\n")
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            parts.append("(empty)")
            continue
        # Trim all-empty trailing rows
        while rows and all(v in (None, "") for v in rows[-1]):
            rows.pop()
        if not rows:
            parts.append("(empty)")
            continue
        max_cols = max(len(r) for r in rows)
        rows = [list(r) + [None] * (max_cols - len(r)) for r in rows]
        # Header row = first non-empty row
        header = rows[0]
        body = rows[1:]
        parts.append("| " + " | ".join(str(c) if c is not None else "" for c in header) + " |")
        parts.append("| " + " | ".join(["---"] * max_cols) + " |")
        for r in body:
            parts.append("| " + " | ".join(
                (str(c) if c is not None else "").replace("\n", " ").replace("|", "\\|")
                for c in r
            ) + " |")
    return "\n".join(parts)


def main() -> int:
    targets = [
        SRC / "Compliance Xai rejection lists.xlsx",
        SRC / "Compliance tracker example.xlsx",
    ]
    for p in targets:
        if not p.exists():
            print(f"missing: {p}")
            continue
        md = xlsx_to_md(p)
        out_name = p.stem.lower().replace(" ", "_") + ".md"
        (OUT / out_name).write_text(md, encoding="utf-8")
        print(f"OK    {p.name} -> {out_name} ({len(md)} chars)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

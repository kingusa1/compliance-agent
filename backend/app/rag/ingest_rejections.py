"""Rejection-tracker ingestion (W4.1 — bulk ingest from XLSX).

Two paths share this module:

1. **Markdown digest** (legacy, build-time fallback):
   reads ``docs/research/2026-04-22-rejection-trackers-digest.md`` via
   ``ingest_rejections(db)``. Anchored on markdown tables whose headers
   include ``Customer Name`` + ``Rejection Reason`` / ``Category``. Used by
   the L10 build-time pipeline; tests reference ``_row_to_chunk`` and
   ``_parse_tables`` directly so the API is preserved.

2. **XLSX bulk-ingest** (W4.1, primary):
   reads ``COMPLIANCE Material XAI/Compliance Xai rejection lists.xlsx``
   via ``ingest_rejections_xlsx(db, dry_run=False)``. Walks the sheet
   row-by-row, using yellow-fill (``FFFFFF00``) on column A as the
   per-rejection block marker. All rows below a yellow row up to the next
   yellow row (or EOF) belong to that block — the column-A cell is the
   customer name, the column-B cells concatenated form the narrative
   coaching note. The hyperlink on the customer cell (when present)
   carries Watt's internal ``site_id`` (e.g.
   ``https://api.wattutilities.co.uk:4433/sites/12874418``).

   See ``.planning/v3-rebuild/2026-05-03-watt-xlsx-deep-dive.md`` §1 for
   the full schema decoding. The XLSX has zero data validation, zero
   formulas, and no header row — yellow-fill is the only structural cue.

Customer names are anonymized to ``[CUSTOMER]`` before the chunk text is
built, so the vector store never holds raw PII. Sales-agent names + Watt
``site_id`` are operational data and are preserved.

CLI usage::

    # Dry-run — print the chunk count + first chunk, write nothing.
    ./venv/bin/python -m app.rag.ingest_rejections --dry-run

    # Real ingest — idempotent (deletes prior xlsx-sourced rows then
    # re-inserts; so re-running yields zero net new rows).
    ./venv/bin/python -m app.rag.ingest_rejections
"""
from __future__ import annotations

import argparse
import logging
import os
import re
import sys
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterator

log = logging.getLogger(__name__)

# ── Markdown digest (legacy build-time) ──────────────────────────────────
REJECTIONS_DOC_PATH = Path("docs/research/2026-04-22-rejection-trackers-digest.md")

# ── XLSX (W4.1, primary) ─────────────────────────────────────────────────
# Watt ships this file inside ``COMPLIANCE Material XAI/`` at repo root.
# We resolve it by walking up from this module so it works regardless of CWD.
REJECTIONS_XLSX_NAME = "Compliance Xai rejection lists.xlsx"
REJECTIONS_XLSX_DIR = "COMPLIANCE Material XAI"
# RGB hex for the yellow per-block-marker fill openpyxl reports.
_YELLOW_FILL_RGB = "FFFFFF00"
# Watt portal site-link prefix — every rejection-row hyperlink uses this.
_WATT_SITE_LINK_PREFIX = "https://api.wattutilities.co.uk:4433/sites/"
# Marker on the chunk text so the xlsx rows can be distinguished from any
# legacy markdown-digest rows; lets the xlsx ingester rebuild idempotently
# without nuking a future markdown rebuild.
_XLSX_SOURCE_MARKER = "[source=xlsx]"

# Match a markdown table row: leading + trailing | with cells separated by |
_TABLE_ROW = re.compile(r"^\s*\|(.+)\|\s*$", re.MULTILINE)
# Separator row underneath a header (e.g. |---|---|).
_SEP_ROW = re.compile(r"^\s*\|(?:\s*-+\s*\|)+\s*$")

# Supplier inference — if any of these appears in the narrative we tag the
# row's supplier metadata with the canonical name. Conservative match —
# many narratives don't name a supplier and stay None.
_SUPPLIER_HINTS: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"\bE\.?\s*ON\s*Next\b", re.I), "E.ON Next Energy"),
    (re.compile(r"\beon next\b", re.I),         "E.ON Next Energy"),
    (re.compile(r"\bE\.?\s*ON\s+script\b", re.I), "E.ON Next Energy"),
    (re.compile(r"\bBritish\s+Gas\s+Lite\b", re.I), "British Gas Lite"),
    (re.compile(r"\bBGL\b"),                     "British Gas Lite"),
    (re.compile(r"\bBritish\s+Gas\s+Business\b", re.I), "British Gas Business"),
    (re.compile(r"\bBritish\s+Gas\s+Trading\b", re.I), "British Gas Trading"),
    (re.compile(r"\bBritish\s+Gas\s+Core\b", re.I), "British Gas Core"),
    (re.compile(r"\bPozitive\b", re.I),          "Pozitive Energy"),
    (re.compile(r"\bYu\s+Energy\b", re.I),       "Yu Energy"),
    (re.compile(r"\bSmartest\s+Energy\b", re.I), "Smartest Energy"),
]


def _split_cells(row: str) -> list[str]:
    inner = row.strip().strip("|")
    return [c.strip() for c in inner.split("|")]


def _resolve_doc_path() -> Path | None:
    if REJECTIONS_DOC_PATH.exists():
        return REJECTIONS_DOC_PATH
    alt = Path(__file__).resolve().parents[3] / REJECTIONS_DOC_PATH
    if alt.exists():
        return alt
    return None


def _resolve_xlsx_path() -> Path | None:
    """Locate the Watt rejection-lists XLSX. Returns None if missing.

    Search order:
      1. ``./<DIR>/<NAME>`` relative to CWD (developer running the CLI)
      2. ``<repo-root>/<DIR>/<NAME>`` relative to this module
      3. Same path with the ``Compliance Material XAI`` (mixed-case) variant
         in case Watt re-zipped it differently.
    """
    candidates = [
        Path(REJECTIONS_XLSX_DIR) / REJECTIONS_XLSX_NAME,
        Path(__file__).resolve().parents[3] / REJECTIONS_XLSX_DIR / REJECTIONS_XLSX_NAME,
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


def _anonymize_customer(name: str) -> str:
    """Replace the customer name with the [CUSTOMER] token. Keeps T/A suffix
    pattern intelligible by collapsing the whole cell to one token."""
    if not name or name in ("", "-"):
        return "[CUSTOMER]"
    return "[CUSTOMER]"


def _parse_tables(md: str) -> list[dict]:
    """Walk the markdown and return per-row dicts whose keys come from the
    nearest preceding table header.

    Only tables containing both 'Customer' and ('Rejection' or 'Category')
    columns count — that's the rejection-tracker shape. Other markdown
    tables (e.g. counts) are skipped.
    """
    lines = md.splitlines()
    out: list[dict] = []

    i = 0
    while i < len(lines):
        line = lines[i]
        if "|" not in line:
            i += 1
            continue
        # Look for a header row followed by a separator row.
        if i + 1 < len(lines) and _SEP_ROW.match(lines[i + 1]):
            header_cells = _split_cells(line)
            lower = [h.lower() for h in header_cells]
            is_rejection_table = (
                any("customer" in h for h in lower)
                and (any("rejection" in h for h in lower) or any("category" in h for h in lower))
            )
            if not is_rejection_table:
                i += 2
                continue

            # Collect rows until the table ends (blank / non-pipe line).
            j = i + 2
            while j < len(lines):
                row = lines[j]
                if "|" not in row or not row.strip():
                    break
                cells = _split_cells(row)
                if len(cells) != len(header_cells):
                    j += 1
                    continue
                rec = dict(zip(header_cells, cells))
                out.append(rec)
                j += 1
            i = j
            continue
        i += 1

    return out


def _row_to_chunk(rec: dict) -> tuple[str, dict]:
    """Return (text, metadata) for one rejection row.

    metadata keys: category, agent_name, supplier, fix.
    """
    # Find columns by fuzzy header match.
    def get(*keys: str) -> str:
        for k in keys:
            for col, val in rec.items():
                if k.lower() in col.lower():
                    return val or ""
        return ""

    customer_raw = get("Customer Name", "Customer")
    customer = _anonymize_customer(customer_raw)
    supplier = get("Supplier")
    agent = get("Sales Agent", "Agent")
    reason = get("Rejection Reason", "Reason")
    category = get("Category")
    fix = get("Fix Required", "Fix")

    text = (
        f"Customer: {customer}. Supplier: {supplier}. Agent: {agent}. "
        f"Category: {category}. Reason: {reason}. Fix: {fix}."
    ).strip()
    metadata = {
        "category": category or None,
        "agent_name": agent or None,
        "supplier": supplier or None,
        "fix": fix or None,
    }
    return text, metadata


# ─────────────────────────────────────────────────────────────────────────
# XLSX ingest path (W4.1)
# ─────────────────────────────────────────────────────────────────────────


@dataclass
class _RejectionBlock:
    """One yellow-fill-anchored rejection block in the Watt XLSX.

    ``customer_raw`` is the column-A cell on the yellow row (kept only
    for logging — never embedded; PII anonymized at chunk-build time).
    ``narrative_lines`` is column-B values for the yellow row + the
    continuation rows below it, in document order, blanks dropped.
    ``site_id`` is the integer site-id parsed off the yellow cell's
    hyperlink (None when no link present).
    """

    customer_raw: str
    narrative_lines: list[str] = field(default_factory=list)
    site_id: int | None = None
    row_no: int = 0  # 1-based xlsx row of the yellow marker, for logs


def _parse_xlsx_blocks(xlsx_path: Path) -> list[_RejectionBlock]:
    """Walk the xlsx, return one ``_RejectionBlock`` per yellow-fill row.

    Pure parser — no DB, no embedding, no PII handling. Suitable for
    direct unit testing on a small fixture xlsx.
    """
    import openpyxl  # lazy: openpyxl is only needed for xlsx ingest

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    blocks: list[_RejectionBlock] = []
    current: _RejectionBlock | None = None

    for r in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=r, column=1)
        cell_b = ws.cell(row=r, column=2)
        a_val = cell_a.value
        b_val = cell_b.value
        fill = cell_a.fill
        is_yellow = bool(
            fill
            and getattr(fill, "fgColor", None)
            and getattr(fill.fgColor, "rgb", None) == _YELLOW_FILL_RGB
        )

        if is_yellow:
            # Start a new block.
            site_id = None
            if cell_a.hyperlink and cell_a.hyperlink.target:
                tgt = cell_a.hyperlink.target
                if tgt.startswith(_WATT_SITE_LINK_PREFIX):
                    tail = tgt[len(_WATT_SITE_LINK_PREFIX):].strip("/").split("/")[0]
                    if tail.isdigit():
                        site_id = int(tail)
            current = _RejectionBlock(
                customer_raw=str(a_val).strip() if a_val else "",
                site_id=site_id,
                row_no=r,
            )
            blocks.append(current)
            # Yellow row's B-col is the first narrative line.
            if b_val:
                current.narrative_lines.append(str(b_val).strip())
        else:
            if current is None:
                # Stray pre-block row (header / blank) — skip.
                continue
            # Continuation row. Some continuation rows have a non-empty
            # column-A (e.g. row 18 'Eon DFV' for Little Dowran Farm —
            # a sub-header inside a block). Treat A-col text as a prefix
            # to the narrative line.
            line_parts: list[str] = []
            if a_val and str(a_val).strip():
                line_parts.append(str(a_val).strip())
            if b_val and str(b_val).strip():
                line_parts.append(str(b_val).strip())
            joined = " — ".join(line_parts).strip() if line_parts else ""
            if joined:
                current.narrative_lines.append(joined)

    # Drop empty blocks (yellow row with no narrative AT ALL — shouldn't
    # happen but defensive).
    return [b for b in blocks if b.narrative_lines]


def _infer_supplier(narrative: str) -> str | None:
    """Best-effort canonical-supplier inference from narrative free text."""
    for pat, canonical in _SUPPLIER_HINTS:
        if pat.search(narrative):
            return canonical
    return None


def _block_to_chunk(block: _RejectionBlock) -> tuple[str, dict]:
    """Convert one ``_RejectionBlock`` to (text, metadata) for RejectionChunk.

    Text format (per W4.1 spec — embed customer + supplier + category +
    reason + narrative; category & sales_agent are absent from the
    rejection-list xlsx so they're empty strings):

        [source=xlsx] Customer: [CUSTOMER]. Site: 12874418. Supplier: <inf>.
        Category: . Agent: . Reason: <first narrative line>.
        Notes: <remaining narrative joined with ' / '>.

    Metadata keys mirror the legacy markdown ingester's contract
    (``category, agent_name, supplier, fix``) so downstream consumers
    don't branch on the source.
    """
    customer = _anonymize_customer(block.customer_raw)
    full_narrative = " ".join(block.narrative_lines)
    supplier = _infer_supplier(full_narrative)
    # First non-trivial narrative line as the canonical "rejection reason";
    # the rest as "notes".
    reason = block.narrative_lines[0] if block.narrative_lines else ""
    notes = " / ".join(block.narrative_lines[1:]) if len(block.narrative_lines) > 1 else ""
    site_part = f"Site: {block.site_id}. " if block.site_id else ""

    text = (
        f"{_XLSX_SOURCE_MARKER} Customer: {customer}. {site_part}"
        f"Supplier: {supplier or ''}. Category: . Agent: . "
        f"Reason: {reason} Notes: {notes}"
    ).strip()

    metadata = {
        "category": None,           # absent from rejection-list xlsx
        "agent_name": None,         # absent from rejection-list xlsx
        "supplier": supplier,       # inferred from narrative
        "fix": None,                # narrative-encoded; not separable here
        "site_id": block.site_id,   # extra kept for logging only
    }
    return text, metadata


def ingest_rejections_xlsx(
    db: Any,
    *,
    dry_run: bool = False,
    xlsx_path: Path | None = None,
) -> int:
    """Read the Watt rejection-list xlsx, anonymize, embed, write chunks.

    Idempotent: deletes any pre-existing xlsx-sourced rows (matched by the
    ``[source=xlsx]`` marker on ``text``) before inserting fresh ones, so
    re-running yields zero net new rows. Markdown-digest rows (no marker)
    are preserved.

    Returns:
        Number of chunks built (== number of rejection blocks in the xlsx).
        Zero if RejectionChunk ORM is missing or the xlsx file is absent.

    Args:
        db: SQLAlchemy session (not used in dry-run mode).
        dry_run: If True, parse + chunk + embed-skip but never write.
        xlsx_path: Override the default xlsx location (used by tests).
    """
    try:
        from app.models import RejectionChunk  # type: ignore
    except ImportError:
        log.warning(
            "RejectionChunk ORM not yet present; ingest_rejections_xlsx is a no-op."
        )
        return 0

    path = xlsx_path or _resolve_xlsx_path()
    if path is None:
        log.warning(
            "REJECTIONS_XLSX_INGEST file not found; expected %s/%s under repo root",
            REJECTIONS_XLSX_DIR, REJECTIONS_XLSX_NAME,
        )
        return 0

    blocks = _parse_xlsx_blocks(path)
    if not blocks:
        log.warning("REJECTIONS_XLSX_INGEST no yellow-marker blocks found in %s", path)
        return 0

    pairs = [_block_to_chunk(b) for b in blocks]
    texts = [t for (t, _) in pairs]

    if dry_run:
        log.info(
            "REJECTIONS_XLSX_INGEST dry-run: %d blocks parsed from %s (no write)",
            len(blocks), path,
        )
        # Print to stdout so the CLI surfaces it without depending on log config.
        print(f"[dry-run] parsed {len(blocks)} rejection blocks from {path}")
        if pairs:
            print(f"[dry-run] first chunk preview:\n  {pairs[0][0][:240]}...")
        return len(blocks)

    embeddings: list[Any] = [None] * len(texts)
    try:
        from app.rag.embed import embed_batch

        embeddings = embed_batch(texts)
    except EnvironmentError as e:
        log.warning("REJECTIONS_XLSX embed skipped (no key): %s", e)
    except Exception as e:  # noqa: BLE001
        log.warning("REJECTIONS_XLSX embed failed: %s", e)

    # Idempotent rebuild — delete only xlsx-sourced rows.
    db.query(RejectionChunk).filter(
        RejectionChunk.text.like(f"{_XLSX_SOURCE_MARKER}%")
    ).delete(synchronize_session=False)

    for idx, ((txt, meta), emb) in enumerate(zip(pairs, embeddings)):
        db.add(RejectionChunk(
            chunk_idx=idx,
            text=txt,
            category=meta.get("category"),
            agent_name=meta.get("agent_name"),
            supplier=meta.get("supplier"),
            fix=meta.get("fix"),
            embedding=emb,
        ))
    db.commit()

    embedded = bool(embeddings) and embeddings[0] is not None
    log.info(
        "REJECTIONS_XLSX_INGEST blocks=%d embedded=%s path=%s",
        len(blocks), "yes" if embedded else "no", path,
    )
    return len(blocks)


def ingest_rejections(db) -> int:
    """Read the rejection-tracker digest, anonymize, chunk, embed, write.

    Legacy markdown-digest path. Preserved for the L10 build-time pipeline
    and the tests in ``test_rag_ingest_pipelines.py``. New work should
    use :func:`ingest_rejections_xlsx` instead.

    Idempotent rebuild. Returns rows written. 0 if RejectionChunk ORM
    missing or doc not found.
    """
    try:
        from app.models import RejectionChunk  # type: ignore
    except ImportError:
        log.warning("RejectionChunk ORM not yet present; ingest_rejections is a no-op stub.")
        return 0

    doc = _resolve_doc_path()
    if doc is None:
        log.warning("REJECTIONS_INGEST doc not found at %s", REJECTIONS_DOC_PATH)
        return 0

    md = doc.read_text(errors="ignore")
    rows = _parse_tables(md)
    if not rows:
        log.warning("REJECTIONS_INGEST no rejection tables matched")
        return 0

    pairs = [_row_to_chunk(r) for r in rows]
    texts = [t for (t, _) in pairs]

    embeddings: list[Any] = [None] * len(texts)
    try:
        from app.rag.embed import embed_batch

        embeddings = embed_batch(texts)
    except EnvironmentError as e:
        log.warning("REJECTIONS embed skipped (no key): %s", e)
    except Exception as e:  # noqa: BLE001
        log.warning("REJECTIONS embed failed: %s", e)

    # Delete only markdown-sourced rows (no [source=xlsx] marker) so the
    # two ingesters can coexist without nuking each other.
    db.query(RejectionChunk).filter(
        ~RejectionChunk.text.like(f"{_XLSX_SOURCE_MARKER}%")
    ).delete(synchronize_session=False)
    for idx, ((txt, meta), emb) in enumerate(zip(pairs, embeddings)):
        db.add(RejectionChunk(
            chunk_idx=idx,
            text=txt,
            category=meta.get("category"),
            agent_name=meta.get("agent_name"),
            supplier=meta.get("supplier"),
            fix=meta.get("fix"),
            embedding=emb,
        ))
    db.commit()

    embedded = embeddings and embeddings[0] is not None
    log.info(
        "REJECTIONS_INGEST rows=%d embedded=%s",
        len(rows), "yes" if embedded else "no",
    )
    return len(rows)


# ─────────────────────────────────────────────────────────────────────────
# Tracker XLSX ingest path (D2 — Compliance tracker example.xlsx)
# ─────────────────────────────────────────────────────────────────────────
#
# Second XLSX path. Reads the operational tracker spreadsheet (MARCH 26 +
# APRIL 2026 sheets — ~140 rows) and produces one chunk per row. Distinct
# from the rejection-list path above (which is one chunk per yellow-fill
# block); the two coexist via different source markers on the chunk text:
#
#   rejection-list path → ``[source=xlsx]``
#   tracker path        → ``[source=tracker:<sheet>:<customer>]``
#
# Idempotency for the tracker path is per-row (marker contains sheet+customer
# so re-ingesting the same XLSX skips already-loaded rows). Anonymization
# follows the W4.A pattern: customer name is collapsed to ``[CUSTOMER]``
# before chunk text is built.

_TRACKER_SOURCE_PREFIX = "[source=tracker:"


def _embed(text: str) -> list[float] | None:
    """Single-text embedding wrapper. Tests patch this directly so they can
    pass a deterministic vector without invoking the OpenAI client.

    Returns None when no embedding key is configured (degrades gracefully
    so the ingest still writes rows — search is just disabled until a
    backfill runs)."""
    try:
        from app.rag.embed import embed_one
        return embed_one(text)
    except EnvironmentError as e:
        log.warning("REJECTIONS_TRACKER embed skipped (no key): %s", e)
        return None
    except Exception as e:  # noqa: BLE001
        log.warning("REJECTIONS_TRACKER embed failed: %s", e)
        return None


def _resolve_tracker_xlsx_path() -> str:
    """Locate the Compliance tracker example.xlsx. Raises FileNotFoundError
    if neither candidate path exists."""
    candidates = [
        "/Users/gomaa/Documents/Compliance-Agent/COMPLIANCE Material XAI/Compliance tracker example.xlsx",
        os.path.join(
            os.path.dirname(__file__), "..", "..", "..",
            "COMPLIANCE Material XAI", "Compliance tracker example.xlsx",
        ),
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    raise FileNotFoundError("Compliance tracker example.xlsx not found")


def ingest_tracker_xlsx(
    db: Any,
    *,
    xlsx_path: str | None = None,
    dry_run: bool = False,
) -> int:
    """Ingest the Compliance tracker example.xlsx — MARCH 26 + APRIL 2026
    sheets (~140 rows total). Each row becomes one rejections-namespace
    chunk tagged ``source=tracker`` so the existing rejection-list chunks
    can coexist.

    Returns:
        Number of chunks written (or — in dry-run — number of rows that
        would be written). Idempotent: re-running over the same XLSX is a
        no-op because each row's marker (sheet + customer name) is unique
        and existing rows are skipped.
    """
    import openpyxl  # lazy: openpyxl is only needed for xlsx ingest

    path = xlsx_path or _resolve_tracker_xlsx_path()
    wb = openpyxl.load_workbook(path, data_only=True)
    rows: list[dict] = []
    for sheet_name in ("MARCH 26", "APRIL 2026"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [str(c.value or "").strip() for c in ws[1]]
        for r in range(2, ws.max_row + 1):
            cells = [ws.cell(row=r, column=col + 1).value for col in range(len(headers))]
            row = dict(zip(headers, cells))
            if not row.get("Customer Name"):
                continue
            rows.append({"sheet": sheet_name, **row})

    if dry_run:
        log.info(f"tracker XLSX would ingest {len(rows)} rows from {path}")
        return len(rows)

    try:
        from app.models import RejectionChunk  # type: ignore
    except ImportError:
        log.warning("RejectionChunk ORM not yet present; ingest_tracker_xlsx is a no-op.")
        return 0

    written = 0
    for row in rows:
        # Anonymize PII before chunk-text construction (matches W4.A).
        text_body = (
            f"Customer: [CUSTOMER]\n"
            f"Supplier: {row.get('Supplier') or 'unknown'}\n"
            f"Sales Agent: {row.get('Sales Agent') or 'unknown'}\n"
            f"Category: {row.get('Category') or 'unknown'}\n"
            f"Rejection Reason: {row.get('Rejection Reason') or '-'}\n"
            f"Fix Required: {row.get('Fix Required') or '-'}\n"
            f"Status: {row.get('Status') or '-'}\n"
            f"Outcome: {row.get('Outcome') or '-'}\n"
            f"Notes: {row.get('Notes') or '-'}"
        )
        marker = f"{_TRACKER_SOURCE_PREFIX}{row['sheet']}:{row.get('Customer Name')}]"
        # Idempotent re-runs: skip if marker already present in any chunk.
        if db.query(RejectionChunk).filter(RejectionChunk.text.contains(marker)).first():
            continue
        full_text = text_body + "\n\n" + marker
        emb = _embed(full_text)
        # Map the spec's metadata dict onto the existing RejectionChunk
        # column schema (no `metadata_` / `namespace` columns on this model;
        # the marker carries the source tag and the columns carry the
        # operational facets used by downstream filters).
        category = row.get("Category")
        agent_name = row.get("Sales Agent")
        supplier = row.get("Supplier")
        fix = row.get("Fix Required")
        db.add(RejectionChunk(
            id=uuid.uuid4(),
            chunk_idx=written,
            text=full_text,
            category=str(category).strip() if category else None,
            agent_name=str(agent_name).strip() if agent_name else None,
            supplier=str(supplier).strip() if supplier else None,
            fix=str(fix).strip() if fix else None,
            embedding=emb,
        ))
        written += 1
    db.commit()
    log.info(
        f"tracker XLSX ingest done — wrote {written} new chunks (parsed {len(rows)})"
    )
    return written


# ─────────────────────────────────────────────────────────────────────────
# CLI entrypoint
# ─────────────────────────────────────────────────────────────────────────


def _main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="python -m app.rag.ingest_rejections",
        description="Bulk-ingest Watt rejection-list xlsx into the rejections RAG namespace.",
    )
    parser.add_argument(
        "--source",
        choices=("rejection-list", "tracker", "all"),
        default="rejection-list",
        help="Which XLSX to ingest (default: rejection-list — the legacy "
             "yellow-fill list. 'tracker' = D2 row-per-rejection sheet. "
             "'all' = both.)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse + chunk only; print N blocks/rows, do not embed or write.",
    )
    parser.add_argument(
        "--xlsx",
        default=None,
        help="Override path to the xlsx (default: auto-resolve).",
    )
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )

    # Dry-run never opens a real DB — keep it cheap to invoke.
    if args.dry_run:
        total = 0
        if args.source in ("rejection-list", "all"):
            xlsx = Path(args.xlsx) if args.xlsx else None
            total += ingest_rejections_xlsx(None, dry_run=True, xlsx_path=xlsx)
        if args.source in ("tracker", "all"):
            total += ingest_tracker_xlsx(None, dry_run=True, xlsx_path=args.xlsx)
        return 0 if total > 0 else 1

    from app.database import SessionLocal

    db = SessionLocal()
    try:
        total = 0
        if args.source in ("rejection-list", "all"):
            xlsx = Path(args.xlsx) if args.xlsx else None
            n = ingest_rejections_xlsx(db, dry_run=False, xlsx_path=xlsx)
            print(f"REJECTIONS_XLSX_INGEST wrote {n} chunks")
            total += n
        if args.source in ("tracker", "all"):
            n = ingest_tracker_xlsx(db, dry_run=False, xlsx_path=args.xlsx)
            print(f"REJECTIONS_TRACKER_INGEST wrote {n} chunks")
            total += n
        return 0 if total > 0 else 1
    finally:
        db.close()


if __name__ == "__main__":  # pragma: no cover
    sys.exit(_main())

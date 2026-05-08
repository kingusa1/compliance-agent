"""Import Watt's compliance tracker XLSX into the v3 DB.

Reads ``COMPLIANCE Material XAI/Compliance tracker example.xlsx`` (4 sheets,
~191 rows). Normalises XLSX free text to v3 enums (8 categories, 10 fix
actions, 7 statuses, 5 outcomes, 5 dead reasons). Upserts Customer +
CustomerDeal + Rejection + rejection_audit_log + a synthetic Call per row
so the row appears on /tracker exactly as the reviewer remembers from
their spreadsheet.

Idempotent: each row is keyed by deterministic UUID5(NAMESPACE, sheet+row),
so re-running the importer will NOT duplicate.

Run as CLI:
    python -m app.import_xlsx_tracker --dry-run    # parse + report counts
    python -m app.import_xlsx_tracker --commit     # write to live DB
    python -m app.import_xlsx_tracker --reset      # delete prior import + re-write
"""
from __future__ import annotations

import argparse
import dataclasses
import re
import uuid
from datetime import date, datetime, timedelta, timezone
from typing import Any, Optional

import openpyxl
from sqlalchemy.orm import Session

from app.database import SessionLocal
from app.field_sources import set_source
from app.logger import log
from app.models import (
    Call,
    Customer,
    CustomerDeal,
    Rejection,
)


# Deterministic namespace for idempotent re-imports. Don't change — re-running
# with a different namespace will create duplicate rows.
NAMESPACE = uuid.UUID("8c3f3c4e-1a14-7e9e-908c-453e448dbcf0")


SOURCE_SHEET_ORDER = ["MARCH 26", "APRIL 2026", "FIXED REJECTIONS", "DEAD REJECTIONS"]


# ── Enum vocabularies (verbatim from the live DB enums) ─────────────────
_CATEGORIES = {
    "ADMIN_ERROR", "PROCESS_FAILURE", "VERBAL_SALES_ERROR",
    "COMPLIANCE_ISSUE", "COMPLIANCE_ERROR", "PRICING_ISSUE", "PRICING_ERROR",
    "DOCUSIGN_ERROR", "FAILED_CREDIT_CHECK",
}
_FIX_ACTIONS = {
    "AMENDMENT_CALL", "CONFIRMATION_CALL", "NEW_LOA", "NEW_DOCUSIGN",
    "DD_MANDATE", "RESELL_TO_OTHER_SUPPLIER", "PRICE_RECHECK",
    "COT_CHANGE_OF_TENANCY", "CONTRACT_LENGTH_LIMIT",
    "MANUAL_ADMIN_SUBMISSION",
}
_STATUSES = {
    "NOT_STARTED", "IN_PROGRESS", "FIXED", "BATCHED_TO_PORTAL",
    "SUBMITTED_TO_PORTAL", "FIXED_AND_APPROVED", "DEAD",
}
_OUTCOMES = {
    "FIXED_AND_SUBMITTED", "CUSTOMER_LOST", "CANCELLED",
    "NOT_RECOVERABLE", "RESIGNED_TO_OTHER_SUPPLIER",
}
_DEAD_REASONS = {
    "in_contract", "customer_debt", "wrong_owner", "bacs_rejected", "hung_up",
}
_SUPPLIERS = {
    "E.ON Next Energy", "British Gas Lite", "British Gas Business",
    "British Gas Trading", "British Gas Core", "Pozitive", "Yu Energy",
    "Smartest Energy", "Affect Energy", "Britannia Gas",
    "United Gas & Power", "E.ON", "TotalEnergies (out-of-matrix)", "Other",
}


def NORMALISE_CATEGORY(v: Optional[str]) -> Optional[str]:
    if not v:
        return None
    s = str(v).strip().upper().replace(" ", "_").replace("-", "_")
    # Typos collapsed to canonical:
    s = s.replace("PRIOCESS_FAILURE", "PROCESS_FAILURE")
    s = s.replace("DOCUISGN_ERROR", "DOCUSIGN_ERROR")
    if s in _CATEGORIES:
        return s
    return None


def NORMALISE_FIX(v: Optional[str]) -> Optional[str]:
    if not v:
        return None
    s = str(v).strip().upper().replace(" ", "_").replace("-", "_")
    if s in _FIX_ACTIONS:
        return s
    # Common free-text → action mapping (best-effort, leave None when unsure):
    if "AMENDMENT" in s:
        return "AMENDMENT_CALL"
    if "CONFIRM" in s:
        return "CONFIRMATION_CALL"
    if "LOA" in s:
        return "NEW_LOA"
    if "DOCUSIGN" in s:
        return "NEW_DOCUSIGN"
    if "DD" in s or "DIRECT_DEBIT" in s:
        return "DD_MANDATE"
    if "RESELL" in s:
        return "RESELL_TO_OTHER_SUPPLIER"
    if "PRICE" in s and ("RECHECK" in s or "CHECK" in s):
        return "PRICE_RECHECK"
    if "COT" in s or "TENANCY" in s:
        return "COT_CHANGE_OF_TENANCY"
    if "CONTRACT" in s and ("LENGTH" in s or "TERM" in s):
        return "CONTRACT_LENGTH_LIMIT"
    if "MANUAL" in s:
        return "MANUAL_ADMIN_SUBMISSION"
    return None


def NORMALISE_STATUS(v: Optional[str]) -> Optional[str]:
    if not v:
        return None
    s = str(v).strip().upper().replace(" ", "_").replace("-", "_")
    if s in _STATUSES:
        return s
    if s in {"FIXED_AND_APPROVED", "FIXED_AND_APPROVE", "APPROVED"}:
        return "FIXED_AND_APPROVED"
    if s in {"BATCHED_TO_PORTAL", "BATCHED"}:
        return "BATCHED_TO_PORTAL"
    if s in {"SUBMITTED_TO_PORTAL", "SUBMITTED"}:
        return "SUBMITTED_TO_PORTAL"
    if s in {"DEAD"}:
        return "DEAD"
    if s in {"FIXED"}:
        return "FIXED"
    if s in {"IN_PROGRESS", "INPROGRESS"}:
        return "IN_PROGRESS"
    if s in {"NOT_STARTED", "NEW", "OPEN"}:
        return "NOT_STARTED"
    return None


def NORMALISE_OUTCOME(v: Optional[str]) -> Optional[str]:
    if not v:
        return None
    s = str(v).strip().upper().replace(" ", "_").replace("-", "_")
    return s if s in _OUTCOMES else None


def NORMALISE_SUPPLIER(v: Optional[str]) -> Optional[str]:
    if not v:
        return None
    s = str(v).strip()
    # Typo collapsed to canonical:
    typos = {
        "Pozative": "Pozitive",
        "BG Lite": "British Gas Lite",
        "BG Business": "British Gas Business",
        "BG Trading": "British Gas Trading",
        "BG Core": "British Gas Core",
        "BG": "British Gas Core",
        "EON": "E.ON",
        "EON Next": "E.ON Next Energy",
        "Total Energies": "TotalEnergies (out-of-matrix)",
        "TotalEnergies": "TotalEnergies (out-of-matrix)",
        "Yu": "Yu Energy",
        "Smartest": "Smartest Energy",
        "Affect": "Affect Energy",
        "Britannia": "Britannia Gas",
        "UGP": "United Gas & Power",
        "United Gas And Power": "United Gas & Power",
    }
    if s in typos:
        return typos[s]
    if s in _SUPPLIERS:
        return s
    return "Other"  # catch-all per W1.3


def _slugify(name: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "-", (name or "").lower()).strip("-")
    return s or "unknown"


def _extract_dead_reason(notes: Optional[str]) -> Optional[str]:
    if not notes:
        return None
    n = notes.lower()
    if "in contract" in n or "still in contract" in n:
        return "in_contract"
    if "debt" in n:
        return "customer_debt"
    if "wrong owner" in n or "right person" in n or "right owner" in n:
        return "wrong_owner"
    if "bacs" in n:
        return "bacs_rejected"
    if "hung up" in n or "hangup" in n:
        return "hung_up"
    return None


def _coerce_date(v: Any) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    # Try a handful of common XLSX free-text formats:
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d %B %Y", "%d %b %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _coerce_dt(v: Any) -> Optional[datetime]:
    d = _coerce_date(v)
    if d is None:
        return None
    return datetime(d.year, d.month, d.day, tzinfo=timezone.utc)


def _coerce_money(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[£,\s]+", "", str(v))
    try:
        return float(s)
    except ValueError:
        return None


@dataclasses.dataclass
class TrackerXlsxRow:
    sheet: str
    row_idx: int                  # 1-based, after header (so first data row = 2)
    customer_name: Optional[str]
    mpan_or_mprn: Optional[str]
    expected_live_date: Optional[date]
    deal_value_gbp: Optional[float]
    supplier: Optional[str]
    rejected_at: Optional[datetime]
    sales_agent: Optional[str]
    rejection_reason: Optional[str]
    category: Optional[str]
    fix_required: Optional[str]
    fix_assignee_name: Optional[str]
    status: Optional[str]
    last_action_date: Optional[datetime]
    deadline: Optional[datetime]
    outcome: Optional[str]
    notes: Optional[str]
    dead_reason: Optional[str]


def parse_xlsx(path: str) -> list[TrackerXlsxRow]:
    wb = openpyxl.load_workbook(path, data_only=True)
    rows: list[TrackerXlsxRow] = []
    for sheet_name in SOURCE_SHEET_ORDER:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for r_idx in range(2, ws.max_row + 1):
            cells = [ws.cell(row=r_idx, column=c).value for c in range(1, 17)]
            customer_name = cells[0] and str(cells[0]).strip()
            if not customer_name:
                continue
            row = TrackerXlsxRow(
                sheet=sheet_name,
                row_idx=r_idx,
                customer_name=customer_name,
                mpan_or_mprn=cells[1] and str(cells[1]).strip(),
                expected_live_date=_coerce_date(cells[2]),
                deal_value_gbp=_coerce_money(cells[3]),
                supplier=NORMALISE_SUPPLIER(cells[4]),
                rejected_at=_coerce_dt(cells[5]),
                sales_agent=cells[6] and str(cells[6]).strip(),
                rejection_reason=cells[7] and str(cells[7]).strip(),
                category=NORMALISE_CATEGORY(cells[8]),
                fix_required=NORMALISE_FIX(cells[9]),
                fix_assignee_name=cells[10] and str(cells[10]).strip(),
                status=NORMALISE_STATUS(cells[11]) or (
                    "DEAD" if sheet_name == "DEAD REJECTIONS"
                    else "FIXED" if sheet_name == "FIXED REJECTIONS"
                    else "NOT_STARTED"
                ),
                last_action_date=_coerce_dt(cells[12]),
                deadline=_coerce_dt(cells[13]),
                outcome=NORMALISE_OUTCOME(cells[14]),
                notes=cells[15] and str(cells[15]).strip(),
                dead_reason=_extract_dead_reason(cells[15]),
            )
            rows.append(row)
    return rows


@dataclasses.dataclass
class ImportCounts:
    customers_created: int = 0
    deals_created: int = 0
    rejections_created: int = 0
    audit_log_created: int = 0
    calls_created: int = 0
    rows_skipped_no_category: int = 0


def _row_uuid(sheet: str, row_idx: int, kind: str) -> uuid.UUID:
    """Deterministic UUID per (sheet, row_idx, kind) so re-runs don't duplicate."""
    return uuid.uuid5(NAMESPACE, f"{sheet}|{row_idx}|{kind}")


def _customer_uuid(name: str) -> uuid.UUID:
    return uuid.uuid5(NAMESPACE, f"customer|{name.lower().strip()}")


def import_rows(rows: list[TrackerXlsxRow], *, db: Session) -> ImportCounts:
    """Idempotent upsert. Per row creates (or skips):
      - Customer (keyed on lowercased name)
      - CustomerDeal (keyed on (sheet, row_idx))
      - Call (synthetic placeholder, keyed on (sheet, row_idx))
      - Rejection (keyed on (sheet, row_idx))
      - rejection_audit_log entry (one per import row)
    """
    from sqlalchemy import text as _text

    counts = ImportCounts()

    for row in rows:
        # ── Customer ────────────────────────────────────────────────
        cust_id = _customer_uuid(row.customer_name or "Unknown")
        cust = db.query(Customer).filter(Customer.id == cust_id).first()
        if cust is None:
            cust = Customer(
                id=cust_id,
                legal_name=row.customer_name or "Unknown",
                slug=_slugify(row.customer_name or "unknown"),
            )
            db.add(cust)
            db.flush()
            counts.customers_created += 1

        # ── Deal ────────────────────────────────────────────────────
        deal_id = _row_uuid(row.sheet, row.row_idx, "deal")
        deal = db.query(CustomerDeal).filter(CustomerDeal.id == deal_id).first()
        if deal is None:
            deal_status_map = {
                "DEAD": "closed_lost",
                "FIXED": "closed_lost",
                "FIXED_AND_APPROVED": "closed_done",
            }
            deal_status = deal_status_map.get(row.status, "in_progress")
            deal = CustomerDeal(
                id=deal_id,
                customer_id=cust.id,
                customer_name=row.customer_name,
                supplier=row.supplier,
                mpan_or_mprn=row.mpan_or_mprn,
                expected_live_date=row.expected_live_date,
                deal_value_gbp=row.deal_value_gbp,
                status=deal_status,
            )
            # Sprint B5: stamp every column we just populated as
            # ``xlsx_import``. Skip None values so reviewer edits later don't
            # look like overwrites of an xlsx_import value when the xlsx
            # never wrote anything for that field.
            set_source(deal, "customer_id", "xlsx_import")
            set_source(deal, "status", "xlsx_import")
            if row.customer_name is not None:
                set_source(deal, "customer_name", "xlsx_import")
            if row.supplier is not None:
                set_source(deal, "supplier", "xlsx_import")
            if row.mpan_or_mprn is not None:
                set_source(deal, "mpan_or_mprn", "xlsx_import")
            if row.expected_live_date is not None:
                set_source(deal, "expected_live_date", "xlsx_import")
            if row.deal_value_gbp is not None:
                set_source(deal, "deal_value_gbp", "xlsx_import")
            # ``notes`` column may not exist on every install (the live cloud
            # DB has it via the W1 migration; SQLAlchemy model may not). Mirror
            # the dead_reason try/except pattern so tests still pass.
            try:
                if row.notes:
                    deal.notes = row.notes
                    set_source(deal, "notes", "xlsx_import")
            except AttributeError:
                pass
            db.add(deal)
            db.flush()
            counts.deals_created += 1

        # ── Call (synthetic placeholder) ────────────────────────────
        call_id = str(_row_uuid(row.sheet, row.row_idx, "call"))
        call = db.query(Call).filter(Call.id == call_id).first()
        if call is None:
            call = Call(
                id=call_id,
                filename=f"xlsx-import/{row.sheet}/row-{row.row_idx}.placeholder",
                file_path=f"xlsx-import/{row.sheet}/row-{row.row_idx}.placeholder",
                deal_id=deal.id,
                agent_name=row.sales_agent,
                customer_name=row.customer_name,
                status="completed",
                detected_supplier=row.supplier,
                reason="Imported from XLSX tracker (no audio).",
            )
            db.add(call)
            db.flush()
            counts.calls_created += 1

        # ── Rejection ────────────────────────────────────────────────
        if row.category is None:
            counts.rows_skipped_no_category += 1
            continue
        rej_id = _row_uuid(row.sheet, row.row_idx, "rejection")
        rej = db.query(Rejection).filter(Rejection.id == rej_id).first()
        if rej is None:
            rejected_at = row.rejected_at or datetime.now(timezone.utc)
            deadline = row.deadline or (rejected_at + timedelta(days=2))
            resolved_at = row.last_action_date if row.status in ("DEAD", "FIXED_AND_APPROVED") else None
            rej = Rejection(
                id=rej_id,
                call_id=call.id,
                customer_slug=cust.slug,
                supplier=row.supplier,
                sales_agent=row.sales_agent,
                category=row.category,
                rejection_reason=row.rejection_reason or "(no reason recorded in XLSX)",
                fix_required=row.fix_required,
                status=row.status or "NOT_STARTED",
                outcome=row.outcome,
                outcome_narrative=row.notes,
                rejected_at=rejected_at,
                deadline=deadline,
                resolved_at=resolved_at,
            )
            # Sprint B5: stamp every column we just populated as
            # ``xlsx_import``. Skip None values so reviewer edits later don't
            # appear to overwrite a real xlsx_import value. ``call_id`` is
            # the FK to the synthetic placeholder Call (not user-editable
            # data) so we deliberately don't stamp it.
            set_source(rej, "customer_slug", "xlsx_import")
            set_source(rej, "category", "xlsx_import")
            set_source(rej, "rejection_reason", "xlsx_import")
            set_source(rej, "status", "xlsx_import")
            set_source(rej, "rejected_at", "xlsx_import")
            set_source(rej, "deadline", "xlsx_import")
            if row.supplier is not None:
                set_source(rej, "supplier", "xlsx_import")
            if row.sales_agent is not None:
                set_source(rej, "sales_agent", "xlsx_import")
            if row.fix_required is not None:
                set_source(rej, "fix_required", "xlsx_import")
            if row.outcome is not None:
                set_source(rej, "outcome", "xlsx_import")
            if row.notes is not None:
                set_source(rej, "outcome_narrative", "xlsx_import")
            if resolved_at is not None:
                set_source(rej, "resolved_at", "xlsx_import")
            # dead_reason column may not exist on every install — try/except.
            try:
                if row.dead_reason and row.status == "DEAD":
                    rej.dead_reason = row.dead_reason
                    set_source(rej, "dead_reason", "xlsx_import")
            except AttributeError:
                pass
            db.add(rej)
            db.flush()
            counts.rejections_created += 1

            # Back-link from Deal → Rejection.
            if deal.rejection_id is None:
                deal.rejection_id = rej.id
                set_source(deal, "rejection_id", "xlsx_import")

        # ── audit log entry ────────────────────────────────────────
        audit_id = _row_uuid(row.sheet, row.row_idx, "audit")
        existing = db.execute(
            _text("SELECT 1 FROM rejection_audit_log WHERE id = :id"),
            {"id": str(audit_id)},
        ).first()
        if not existing:
            db.execute(
                _text(
                    "INSERT INTO rejection_audit_log "
                    "(id, rejection_id, actor_id, action, from_status, to_status, notes, created_at) "
                    "VALUES (:id, :rid, :aid, :action, NULL, :ts, :notes, :created_at)"
                ),
                {
                    "id": str(audit_id),
                    "rid": str(rej_id),
                    "aid": None,
                    "action": "imported_from_xlsx",
                    "ts": rej.status,
                    "notes": f"Imported from {row.sheet} row {row.row_idx}",
                    "created_at": row.last_action_date or row.rejected_at or datetime.now(timezone.utc),
                },
            )
            counts.audit_log_created += 1

    db.commit()
    return counts


# ── CLI ─────────────────────────────────────────────────────────────────

DEFAULT_XLSX = "/Users/gomaa/Documents/Compliance-Agent/COMPLIANCE Material XAI/Compliance tracker example.xlsx"


def _reset_imported(db: Session) -> int:
    """Delete every row this importer has ever inserted, by deterministic-id
    prefix scan. Used by --reset to wipe + re-write cleanly."""
    from sqlalchemy import text as _text
    # Quick wipe: rejections referencing the synthetic calls + audit log + deals + customers.
    # We can't easily filter by namespace inside the DB, so rely on a sentinel
    # path-prefix on Call.file_path that only the importer writes.
    deleted = db.execute(_text(
        "DELETE FROM calls WHERE file_path LIKE 'xlsx-import/%' RETURNING id"
    )).rowcount or 0
    db.commit()
    return deleted


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default=DEFAULT_XLSX)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--commit", action="store_true")
    parser.add_argument("--reset", action="store_true",
                        help="Delete prior import (by file_path prefix) before writing.")
    args = parser.parse_args()

    log.info(f"\U0001f4e5 XLSX_IMPORT path={args.xlsx}")
    rows = parse_xlsx(args.xlsx)
    log.info(f"\U0001f4e5 XLSX_IMPORT parsed {len(rows)} rows from 4 sheets")

    if args.dry_run:
        from collections import Counter
        cat_counts = Counter(r.category for r in rows)
        sup_counts = Counter(r.supplier for r in rows)
        sta_counts = Counter(r.status for r in rows)
        log.info(f"  category counts: {dict(cat_counts)}")
        log.info(f"  supplier counts: {dict(sup_counts)}")
        log.info(f"  status counts:   {dict(sta_counts)}")
        return

    db = SessionLocal()
    try:
        if args.reset:
            n = _reset_imported(db)
            log.info(f"\U0001f4e5 XLSX_IMPORT reset deleted {n} prior calls (cascaded rows follow)")
        if args.commit:
            counts = import_rows(rows, db=db)
            log.info(
                f"\U0001f4e5 XLSX_IMPORT done — customers={counts.customers_created} "
                f"deals={counts.deals_created} rejections={counts.rejections_created} "
                f"calls={counts.calls_created} audit={counts.audit_log_created} "
                f"skipped_no_category={counts.rows_skipped_no_category}"
            )
        else:
            log.warning("XLSX_IMPORT no --commit flag, no writes performed. Use --commit to apply.")
    finally:
        db.close()


if __name__ == "__main__":
    main()

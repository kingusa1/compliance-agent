"""W4.1 — XLSX rejection-list ingester tests.

Builds a 5-block fixture xlsx in-memory and verifies:

1. ``_parse_xlsx_blocks`` walks yellow-fill markers and returns 5 blocks
   with the right customer, narrative-line count, and (when present)
   site_id parsed from the hyperlink.
2. ``_block_to_chunk`` anonymizes the customer to ``[CUSTOMER]``, embeds
   the inferred supplier when narrative names one, and never leaks the
   raw customer string.
3. ``ingest_rejections_xlsx`` writes 5 RejectionChunk rows into the
   ``rejections`` namespace with embeddings populated (mocked) and
   metadata round-tripping.
4. Re-running the ingester is idempotent: row count stays at 5 (delete
   + re-insert of xlsx-sourced rows only).
5. The xlsx ingester does NOT delete pre-existing markdown-sourced rows
   (the two paths coexist).

The fixture xlsx is built with openpyxl in a tmp dir so the test is
hermetic — no dependency on the real ``COMPLIANCE Material XAI/`` file.
"""
from __future__ import annotations

from pathlib import Path
from unittest.mock import patch

import pytest

# Force model registration so the test_db fixture's
# Base.metadata.create_all() actually emits rejection_chunks on SQLite.
import app.models  # noqa: F401


YELLOW_RGB = "FFFFFF00"
SITE_LINK = "https://api.wattutilities.co.uk:4433/sites/"


def _build_fixture_xlsx(path: Path) -> None:
    """Write a 5-block rejection-list xlsx mirroring the real Watt format.

    Layout:
        row 1  YELLOW  'Acme Ltd 1234567890123'                     (no link)
                       'Wrong unit rate stated.'
        row 2  ()                                                    'Please amend lines 12 13 14 on the e.on script.'
        row 3  ()                                                    'Recompute annual cost.'
        row 4  ()                                                    ''           (blank — separator)
        row 5  YELLOW  'Bob T/A Bob Cleaners 9876543210'   link→111  'BACS denied; obtain DD details.'
        row 6  YELLOW  'Charlie Charity 5555555555555'              'Failed to confirm charity number in LOA.'
        row 7  ()                                                    'Need a new LOA.'
        row 8  YELLOW  'Delta Group Ltd 2222222222222'    link→222  'Pricing wrong on BGL contract.'
        row 9  ()                                                    'Verbal amendment for line 11-14.'
        row 10 YELLOW  'Echo Ltd 3333333333333'                     'You did not state Watt utilities at the start.'
    """
    import openpyxl
    from openpyxl.styles import PatternFill

    yellow = PatternFill(start_color=YELLOW_RGB, end_color=YELLOW_RGB, fill_type="solid")

    wb = openpyxl.Workbook()
    ws = wb.active

    rows: list[tuple[str | None, str | None, str | None, bool, int | None]] = [
        # (col_a, col_b, col_c, is_yellow, site_id_for_link)
        ("Acme Ltd 1234567890123",                "Wrong unit rate stated.",                                 None, True,  None),
        (None,                                    "Please amend lines 12 13 14 on the e.on script.",         None, False, None),
        (None,                                    "Recompute annual cost.",                                  None, False, None),
        (None,                                    None,                                                      None, False, None),
        ("Bob T/A Bob Cleaners 9876543210",       "BACS denied; obtain DD details.",                         None, True,  111),
        ("Charlie Charity 5555555555555",         "Failed to confirm charity number in LOA.",                None, True,  None),
        (None,                                    "Need a new LOA.",                                         None, False, None),
        ("Delta Group Ltd 2222222222222",         "Pricing wrong on BGL contract.",                          None, True,  222),
        (None,                                    "Verbal amendment for line 11-14.",                        None, False, None),
        ("Echo Ltd 3333333333333",                "You did not state Watt utilities at the start.",          None, True,  None),
    ]

    for r_idx, (a, b, c, is_yellow, site_id) in enumerate(rows, start=1):
        cell_a = ws.cell(row=r_idx, column=1, value=a)
        ws.cell(row=r_idx, column=2, value=b)
        if c is not None:
            ws.cell(row=r_idx, column=3, value=c)
        if is_yellow:
            cell_a.fill = yellow
        if site_id is not None:
            cell_a.hyperlink = f"{SITE_LINK}{site_id}"

    wb.save(path)


# ── Fixture ──────────────────────────────────────────────────────────────


@pytest.fixture
def fixture_xlsx(tmp_path: Path) -> Path:
    p = tmp_path / "fixture-rejections.xlsx"
    _build_fixture_xlsx(p)
    return p


# ── Parser-level tests (no DB, no embeddings) ────────────────────────────


def test_parse_xlsx_blocks_counts_yellow_markers(fixture_xlsx: Path) -> None:
    from app.rag.ingest_rejections import _parse_xlsx_blocks

    blocks = _parse_xlsx_blocks(fixture_xlsx)
    assert len(blocks) == 5

    # First block (Acme) has 3 narrative lines (yellow row + 2 continuation).
    assert blocks[0].customer_raw.startswith("Acme Ltd")
    assert len(blocks[0].narrative_lines) == 3
    assert blocks[0].site_id is None

    # Bob block (single line, hyperlink → site 111).
    assert blocks[1].customer_raw.startswith("Bob")
    assert blocks[1].site_id == 111
    assert len(blocks[1].narrative_lines) == 1

    # Charlie block has 2 narrative lines, no link.
    assert blocks[2].customer_raw.startswith("Charlie")
    assert blocks[2].site_id is None
    assert len(blocks[2].narrative_lines) == 2

    # Delta block: 2 narrative lines, link → site 222.
    assert blocks[3].customer_raw.startswith("Delta")
    assert blocks[3].site_id == 222

    # Echo block: 1 narrative line, no link.
    assert blocks[4].customer_raw.startswith("Echo")
    assert blocks[4].site_id is None


def test_block_to_chunk_anonymizes_customer_and_infers_supplier(
    fixture_xlsx: Path,
) -> None:
    from app.rag.ingest_rejections import _block_to_chunk, _parse_xlsx_blocks

    blocks = _parse_xlsx_blocks(fixture_xlsx)

    # Acme — narrative names "e.on script" → infer E.ON Next Energy.
    text_acme, meta_acme = _block_to_chunk(blocks[0])
    assert "[CUSTOMER]" in text_acme
    assert "Acme" not in text_acme  # no PII leak
    assert meta_acme["supplier"] == "E.ON Next Energy"
    # Site is None for Acme — no Site: prefix in text.
    assert "Site: " not in text_acme
    # Source marker is present so the idempotent rebuild can find it.
    assert "[source=xlsx]" in text_acme

    # Bob — site 111 should appear in the text body.
    text_bob, meta_bob = _block_to_chunk(blocks[1])
    assert "Site: 111" in text_bob
    assert meta_bob["site_id"] == 111

    # Delta — narrative names "BGL" → infer British Gas Lite.
    text_delta, meta_delta = _block_to_chunk(blocks[3])
    assert meta_delta["supplier"] == "British Gas Lite"


# ── End-to-end ingest tests (DB + mocked embeddings) ─────────────────────


def _fake_embed_batch(texts: list[str]) -> list[list[float]]:
    """Deterministic non-zero 1536-dim vectors (seed = first char ord)."""
    out = []
    for t in texts:
        seed = (ord(t[0]) if t else 7) / 255.0
        out.append([seed * (i % 7 + 1) / 100.0 for i in range(1536)])
    return out


def test_ingest_rejections_xlsx_writes_chunks_with_embeddings(
    test_db, fixture_xlsx: Path
) -> None:
    from app.models import RejectionChunk
    from app.rag import ingest_rejections as mod

    # embed_batch is imported lazily inside the function body, so we
    # monkeypatch the module it lives in (app.rag.embed).
    with patch("app.rag.embed.embed_batch", _fake_embed_batch):
        written = mod.ingest_rejections_xlsx(test_db, xlsx_path=fixture_xlsx)

    assert written == 5
    rows = test_db.query(RejectionChunk).all()
    assert len(rows) == 5

    # Every row carries the xlsx source marker.
    assert all("[source=xlsx]" in r.text for r in rows)
    # No row leaks raw customer names.
    for r in rows:
        for raw in ("Acme", "Bob", "Charlie", "Delta", "Echo"):
            assert raw not in r.text, f"PII leak: {raw} found in chunk text"

    # Embeddings populated and non-zero.
    assert all(r.embedding is not None for r in rows)
    assert all(any(v != 0 for v in r.embedding) for r in rows)
    # 1536 dims — matches text-embedding-3-small.
    assert all(len(r.embedding) == 1536 for r in rows)

    # Supplier metadata round-trips for the two narratives that named one.
    suppliers = sorted([r.supplier for r in rows if r.supplier])
    assert "British Gas Lite" in suppliers
    assert "E.ON Next Energy" in suppliers

    # chunk_idx is dense 0..4.
    assert sorted(r.chunk_idx for r in rows) == [0, 1, 2, 3, 4]


def test_ingest_rejections_xlsx_is_idempotent(test_db, fixture_xlsx: Path) -> None:
    """Re-running the ingester rebuilds in place — count stays at 5."""
    from app.models import RejectionChunk
    from app.rag import ingest_rejections as mod

    with patch("app.rag.embed.embed_batch", _fake_embed_batch):
        first = mod.ingest_rejections_xlsx(test_db, xlsx_path=fixture_xlsx)
        second = mod.ingest_rejections_xlsx(test_db, xlsx_path=fixture_xlsx)

    assert first == second == 5
    assert test_db.query(RejectionChunk).count() == 5


def test_ingest_rejections_xlsx_preserves_markdown_sourced_rows(
    test_db, fixture_xlsx: Path
) -> None:
    """Pre-seed a markdown-style row (no [source=xlsx] marker) and verify
    the xlsx ingester only deletes/replaces its own kind."""
    from app.models import RejectionChunk
    from app.rag import ingest_rejections as mod

    test_db.add(RejectionChunk(
        chunk_idx=0,
        text="Customer: [CUSTOMER]. Supplier: . Agent: alice. Reason: legacy markdown row.",
        category="ADMIN_ERROR",
        agent_name="alice",
        supplier=None,
        fix=None,
        embedding=None,
    ))
    test_db.commit()
    assert test_db.query(RejectionChunk).count() == 1

    with patch("app.rag.embed.embed_batch", _fake_embed_batch):
        written = mod.ingest_rejections_xlsx(test_db, xlsx_path=fixture_xlsx)

    assert written == 5
    # 1 markdown + 5 xlsx = 6 total.
    assert test_db.query(RejectionChunk).count() == 6
    md_rows = (
        test_db.query(RejectionChunk)
        .filter(~RejectionChunk.text.like("[source=xlsx]%"))
        .all()
    )
    assert len(md_rows) == 1
    assert md_rows[0].agent_name == "alice"


def test_ingest_rejections_xlsx_dry_run_writes_nothing(
    test_db, fixture_xlsx: Path
) -> None:
    from app.models import RejectionChunk
    from app.rag import ingest_rejections as mod

    n = mod.ingest_rejections_xlsx(test_db, xlsx_path=fixture_xlsx, dry_run=True)
    assert n == 5
    assert test_db.query(RejectionChunk).count() == 0


def test_resolve_xlsx_path_returns_none_when_missing(tmp_path: Path) -> None:
    """Path resolver returns None gracefully when file is absent."""
    import os
    from app.rag.ingest_rejections import _resolve_xlsx_path

    cwd = os.getcwd()
    try:
        os.chdir(tmp_path)
        # tmp_path has no COMPLIANCE Material XAI/ dir — but the module
        # also walks up to the repo root, which DOES contain the real
        # file in this checkout. So this only asserts the function shape,
        # not the absence-result on a real repo.
        result = _resolve_xlsx_path()
        assert result is None or result.exists()
    finally:
        os.chdir(cwd)

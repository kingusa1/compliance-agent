"""build_xlsx returns bytes with 4 sheets matching Watt's source schema."""
from __future__ import annotations

import io
import uuid
from datetime import datetime, timedelta, UTC

import openpyxl

from app.tracker_export import build_xlsx
from app.models import Call, Customer, CustomerDeal, Rejection


def test_xlsx_has_4_sheets_with_correct_headers(test_db):
    cust = Customer(id=uuid.uuid4(), legal_name="Acme", slug="acme")
    deal = CustomerDeal(
        id=uuid.uuid4(), customer_id=cust.id,
        customer_name="Acme", supplier="E.ON Next",
        status="in_progress",
    )
    call = Call(
        id=str(uuid.uuid4()), filename="t.mp3", file_path="/tmp/t.mp3",
        deal_id=deal.id, status="completed",
    )
    rej = Rejection(
        id=uuid.uuid4(), call_id=call.id,
        category="VERBAL_SALES_ERROR",
        rejection_reason="Missed disclosure",
        fix_required="AMENDMENT_CALL", status="NOT_STARTED",
        rejected_at=datetime.now(UTC),
        deadline=datetime.now(UTC) + timedelta(days=2),
    )
    test_db.add_all([cust, deal, call, rej])
    test_db.commit()

    data = build_xlsx(test_db)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["MARCH 26", "APRIL 2026", "FIXED REJECTIONS", "DEAD REJECTIONS"]

    expected_headers = [
        "Customer Name", "MPAN / MPRN", "Expected Live date ", "Deal Value (£)",
        "Supplier", "Rejected at", "Sales Agent", "Rejection Reason", "Category",
        "Fix Required", "Fixed BY ", "Status", "Last Action Date", "Deadline",
        "Outcome", "Notes",
    ]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [(c.value or "") for c in ws[1][:16]]
        # Whitespace preserved verbatim — exact match incl trailing spaces.
        assert headers == expected_headers, \
            f"sheet {sheet_name} headers mismatch: {headers!r}"
